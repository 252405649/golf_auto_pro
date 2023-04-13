"""
==========================
Author:sunk123
Time:2022-07-14 20:30
Company:力石科技有限公司
==========================
"""
import openpyxl

class HandleExcel:
    def __init__(self, filename, sheetname):
        # filename   excel文件名称（路径）
        # sheetname    表单名称
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        #获取exce数据
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        # 获取第一行的表头
        title = [i.value for i in res[0]]
        cases = []
        # 遍历第一行之外的其他行
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        # 返回读取出来的数据
        return cases

    # 数据写入的方法
    def write_data(self, row, column, value):
        # 加载工作薄对象
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        # 写入数据
        sh.cell(row=row,column=column,value=value)
        workbook.save(self.filename)

if __name__ == '__main__':
    excel = HandleExcel(r'D:\自动化视频\接口自动化视频\用例数据要求基础模板.xlsx', 'register')
    res = excel.read_data()
    print(res)